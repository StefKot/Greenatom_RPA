import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import re
import openpyxl
from datetime import date
from openpyxl.styles import Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# Получаем текущую дату и форматируем ее
today = date.today()
formatted_date = today.strftime("%d-%m-%Y")

#Запускаем Браузер
driver = webdriver.Chrome()
print("Начинаем работу парсера")

# URL сайта MOEX для валютных курсов
url = 'https://www.moex.com/ru/derivatives/currency-rate.aspx?currency=USD_RUB#'


#Открываем браузер на весь экран
driver.get(url)
get_url = driver.current_url
wait = WebDriverWait(driver, 20) # Ожидание до 20 секунд
print("Current url:", get_url)
driver.maximize_window() # Развернуть браузер на весь экран


# Принимаем соглашение на сайте
agree = wait.until(EC.element_to_be_clickable(
    (By.XPATH, "//*[contains(text(), 'Согласен')]")))
agree.click()

# USD
# Получаем URL текущей страницы
get_url = driver.current_url

# Селектор для таблицы с курсами
table = '//*[@id="app"]/div[1]/div[2]/div/div[2]/div[3]/table'
table_1 = wait.until(EC.visibility_of_element_located((By.XPATH, table)))
table_1_value = table_1.text # Получаем текст таблицы

# Формируем имя файла Excel
excel_file = f"result_{formatted_date}.xlsx"

# Открываем файл Excel
workbook = openpyxl.Workbook()
# Получаем активный лист
worksheet = workbook.active

# Задаем заголовки столбцов
worksheet['A1'] = 'Дата USD/RUB'
worksheet['B1'] = 'Курс USD/RUB'
worksheet['C1'] = 'Время USD/RUB'
worksheet['G1'] = 'Результат'


# Разделяем строки таблицы и записываем данные в файл Excel
table_rows = table_1_value.replace('\r','').split('\n')
for row in range(7, len(table_rows)):
    # Задаём ячейкам финансовый формат
    table_cols = table_rows[row].split()
    date = table_cols[0]
    value = table_cols[3]
    time = table_cols[2]
    # print(table_cols)
     # Задаем формат для ячеек с курсом
    worksheet.cell(row=row, column=2, value=value).number_format = '# ##0.0000' + ' ' + u'\u20BD'
    worksheet.cell(row=row, column=5, value=value).number_format = '# ##0.0000' + ' ' + u'\u20BD'
    worksheet.cell(row=row, column=7).number_format = '# ##0.0000' + ' ' + u'\u20BD'

    # Заполняем колонки
    value_num = float(value)  # Преобразуем строку в число
    worksheet.cell(row=row, column=1, value=date)
    worksheet.cell(row=row, column=2, value=value_num)
    worksheet.cell(row=row, column=3, value=time)


# Сохраняем файл
workbook.save(excel_file)




# JPY
# Получаем URL для JPY/RUB
get_url = re.sub(r'currency=([^/]*)', r'currency=JPY_RUB', get_url)
print(f"url для JPY/RUB: {get_url}")
# Переходим на страницу с курсом JPY/RUB
driver.get(get_url)

# Получаем текст таблицы с курсами JPY/RUB
table_2 = wait.until(EC.visibility_of_element_located((By.XPATH, table)))
table_2_value = table_2.text
driver.close()

# Открываем файл Excel
workbook = openpyxl.load_workbook(excel_file)
# Получаем активный лист
worksheet = workbook.active

# Задаем заголовки столбцов для JPY/RUB
worksheet['D1'] = 'Дата JPY/RUB'
worksheet['E1'] = 'Курс JPY/RUB'
worksheet['F1'] = 'Время JPY/RUB'

# Разделяем строки таблицы и записываем данные в файл Excel
table_rows = table_2_value.replace('\r','').split('\n')
for row in range(7, len(table_rows)):
    table_cols = table_rows[row].split()
    date = table_cols[0]
    value = table_cols[3]
    time = table_cols[2]
    value_num = float(value)
    # Заполняем колонки для JPY/RUB
    worksheet.cell(row=row, column=4, value=date)
    worksheet.cell(row=row, column=5, value=value_num)
    worksheet.cell(row=row, column=6, value=time)


# Выравнивание содержимого ячеек по центру
for row in worksheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Автоширина столбцов
for col in worksheet.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    worksheet.column_dimensions[col_letter].width = adjusted_width


worksheet = workbook.active

# Удаляем пустые строки
max_row = worksheet.max_row
for row in range(max_row, 1, -1):
    if not any(cell.value for cell in worksheet[row]):
        worksheet.delete_rows(row)

# Вычисляем результат деления курса USD/RUB на курс JPY/RUB
num_rows = worksheet.max_row
for row in range(2, num_rows + 1):
    worksheet.cell(row=row, column=7).value = f"=B{row}/E{row}"
print(num_rows)

# Сохраняем файл
workbook.save(excel_file)

# Функция склонения числительных
forms = ['строка', 'строки', 'строк']
def declension(num, forms):
    if num % 100 in [11, 12, 13, 14]:
        return f'{num} {forms[2]}'
    elif num % 10 == 1:
        return f'{num} {forms[0]}'
    elif num % 10 in [2, 3, 4]:
        return f'{num} {forms[1]}'
    else:
        return f'{num} {forms[2]}'

# Функция отправки email
def send_mail():
    # Указываем параметры для подключения к серверу SMTP
    smtp_host = 'smtp.mail.ru' # SMTP-сервер
    smtp_port = 587 # Порт SMTP-сервера
    smtp_user = 'adac-parser@mail.ru' # Логин
    smtp_password = 'B2KjWWX6ma9sq4YhXDEv' # Пароль

    # Создаем объект MIMEMultipart для сообщения
    msg = MIMEMultipart()
    msg['From'] = 'adac-parser@mail.ru'
    msg['To'] = 'adac-parser@mail.ru'
    msg['Subject'] = 'Поддержка RPA'

    # Добавляем текст сообщения
    body = text
    msg.attach(MIMEText(body, 'plain'))

    # Добавляем вложение (файл Excel)
    with open(f'{excel_file}', 'rb') as f:
        attach = MIMEApplication(f.read(), _subtype='xlsx')
        attach.add_header('Content-Disposition', 'attachment', filename=f'{excel_file}')
        msg.attach(attach)

    # Отправляем письмо
    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)

    print("\033[92mПисьмо отправлено\033[0m")

# Формируем текст сообщения
text = f"В таблице {declension(num_rows, forms)}"
declension(num_rows, forms)
# Выводим информацию о количестве строк в таблице
print(text)
send_mail()
